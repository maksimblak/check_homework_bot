from enum import Enum  # Импорт модуля Enum для создания перечислений.
import openpyxl  # Импорт модуля openpyxl для работы с Excel-файлами.
from openpyxl.workbook.workbook import Workbook  # Импорт класса Workbook из openpyxl.
from openpyxl.worksheet.worksheet import Worksheet  # Импорт класса Worksheet из openpyxl.
from model.pydantic.db_start_data import StudentRaw, TeacherRaw


class ExcelDataParserError(Exception):
    """Пользовательский класс исключения для ошибок при парсинге данных из Excel."""

    def __init__(self, message):
        super().__init__(message)


class ParserType(Enum):  # Создание перечисления ParserType для определения типа парсинга.
    ALL = 0
    TEACHER = 1
    STUDENT = 2


class ExcelDataParser:  # Класс для парсинга данных из Excel-файлов.

    def __init__(self, file_path: str, parse_type: ParserType = ParserType.ALL):
        # Инициализация класса с заданием типа парсинга по умолчанию.
        self.__student: dict[str, dict[str, list[StudentRaw]]] = {}  # Словарь для данных студентов.
        self.__teacher: dict[str, dict[str, list[TeacherRaw]]] = {}  # Словарь для данных учителей.
        self.__parse_type = parse_type  # Задание типа парсинга.
        self.__load_data(file_path, parse_type)  # Загрузка данных.

    @property  # Свойство для доступа к данным студентов, с учетом типа парсинга.
    def students(self) -> dict[str, dict[str, list[StudentRaw]]]:
        if self.__parse_type == ParserType.TEACHER:
            raise ExcelDataParserError("Students data don't with this ParseType")
        return self.__student

    @property  # Свойство для доступа к данным учителей, с учетом типа парсинга.
    def teachers(self) -> dict[str, dict[str, list[TeacherRaw]]]:
        if self.__parse_type == ParserType.STUDENT:
            raise ExcelDataParserError("Teachers data don't with this ParseType")
        return self.__teacher

    def __load_data(self, file_path: str, parse_type: ParserType) -> None:
        wb: Workbook = openpyxl.load_workbook(file_path)  # Загрузка Excel-файла в объект Workbook.

        match parse_type:

            case ParserType.ALL:
                index = wb.sheetnames.index('teachers')  # Находим индекс листа "teachers".
                wb.active = index  # Устанавливаем активный лист.
                self.__teachers_parser(wb.active)  # Вызываем метод для парсинга данных учителей.

                index = wb.sheetnames.index('students')  # Находим индекс листа "students".
                wb.active = index  # Устанавливаем активный лист.
                self.__students_parser(wb.active)  # Вызываем метод для парсинга данных студентов.

            case ParserType.STUDENT:
                index = wb.sheetnames.index('students')  # Находим индекс листа "students" из Excel-файла.
                wb.active = index  # Устанавливаем активный лист.
                self.__students_parser(wb.active)  # Вызываем метод для парсинга данных о студентах.

            case ParserType.TEACHER:
                index = wb.sheetnames.index('teachers')  # Находим индекс листа "teachers".
                wb.active = index  # Устанавливаем активный лист.
                self.__teachers_parser(wb.active)  # Вызываем метод для парсинга данных учителей.

            case _:
                raise ExcelDataParserError('ParserType not found')
                # Генерируем исключение при неверном типе парсинга (ошибочно передали информацию)

    def __teachers_parser(self, worksheet: Worksheet) -> None:  # Метод для парсинга данных учителей из листа Excel.
        teacher_name = ''  # Инициализация имени учителя.
        row = 2  # Начальная строка для чтения данных.

        while teacher_name is not None:
            teacher_name = worksheet.cell(row=row, column=1).value  # Извлечение имени учителя из ячейки.
            telegram_id = worksheet.cell(row=row, column=2).value  # Извлечение Telegram ID из ячейки.
            discipline = worksheet.cell(row=row, column=3).value  # Извлечение дисциплины из ячейки.
            is_admin = bool(worksheet.cell(row=row, column=4).value)  # Извлечение информации об администраторе.
            group = worksheet.cell(row=row, column=5).value  # Извлечение группы из ячейки.

            if teacher_name is None:
                break

            if discipline not in self.__teacher:
                self.__teacher[discipline] = {}  # Создание словаря для дисциплины, если еще не создан.

            if group not in self.__teacher[discipline]:
                self.__teacher[discipline][group] = []  # Создание списка для группы, если еще не создан.

            self.__teacher[discipline][group].append(
                TeacherRaw(full_name=teacher_name, telegram_id=telegram_id, is_admin=is_admin))
            row += 1

    def __students_parser(self, worksheet: Worksheet) -> None:  # Метод для парсинга данных студентов из листа Excel.
        student_name = ''  # Инициализация имени студента.
        row = 2  # Начальная строка для чтения данных.

        while student_name is not None:
            student_name = worksheet.cell(row=row, column=1).value  # Извлечение имени студента из ячейки.
            group = worksheet.cell(row=row, column=2).value  # Извлечение группы из ячейки.
            discipline = worksheet.cell(row=row, column=3).value  # Извлечение дисциплины из ячейки.

            if student_name is None:
                break

            if discipline not in self.__student:
                self.__student[discipline] = {}  # Создание словаря для дисциплины, если еще не создан.

            if group not in self.__student[discipline]:
                self.__student[discipline][group] = []  # Создание списка для группы, если еще не создан.

            self.__student[discipline][group].append(
                StudentRaw(full_name=student_name))

            row += 1
